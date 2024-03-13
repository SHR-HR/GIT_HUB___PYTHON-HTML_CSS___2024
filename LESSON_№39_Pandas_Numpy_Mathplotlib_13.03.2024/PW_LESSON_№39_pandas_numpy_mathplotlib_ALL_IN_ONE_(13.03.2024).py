# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
"""""
Дата выполнения ПРАКТИЧЕСКОЙ РАБОТЫ: 14 МАРТА 2024 года.
"""""
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
                                            Практическая работа

                                            Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
                                                                         Дисциплина: Основы программирования на Python

                Практическая работа №39: Работа с большими данными - pandas, numpy, mathplotlib


                                            Выполните следующие задания:


                    Задание №1
                    а) Получите матрицу данных с рандомно заполненного листа с excel – файла.
                    б) «Загрузите» данные с матрицы в numpy массивы.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Урок от 13.03.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Вариант №1.
'''
import pandas as pd
import numpy as np

# Мне было лень создавать отдельно Excel файл, "рисовать" там таблицу и выдумывать самостоятельно какие-то рандомные,
# данные или числа и т.д. Поэтому я сделал это все через Python.

data = np.random.rand(5, 5)  # Для начала создаем матрицу 5x5 с рандомными значениями от 0 до 1

# Затем нам необходимо сделать преобразование матрицы в DataFrame
df = pd.DataFrame(data, columns=[f'Column_{i+1}' for i in range(data.shape[1])])

# После Мы сохраняем полученные данные в Excel-файл
excel_file = 'random_data.xlsx'  # Просто рандомное имя Excel-файла
df.to_excel(excel_file, index=False)  # Тут я подумал, что лучше сохранить данные без индексов

# Теперь Мы сможем с легкостью получить матрицу данных с рандомно заполненного листа Excel-файла
df_from_excel = pd.read_excel(excel_file)

# Далее нам нужно совершить загрузку (считывание) данных с матрицы в numpy массивы
numpy_array = df_from_excel.to_numpy()

# Ну и под конец, необходимо сделать вывод первых нескольких строк массива для проверки
print(numpy_array[:5])  # Печать первых 5 строк массива (можно изменить число строк)

# Теперь у Нас есть Excel-файл с рандомными данными и данные загружены в NumPy массивы

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Импортируем библиотеки pandas и numpy:
'''
import pandas as pd
import numpy as np
'''
pandas используется для работы с данными в виде таблиц (DataFrame),
а numpy - для работы с многомерными массивами. 
Импортируются как pd и np соответственно для удобства обращения к ним.
'''

'''
Создаем случайную матрицу размером 5x5, заполненную случайными
значениями от 0 до 1 с помощью функции numpy.random.rand():
'''
data = np.random.rand(5, 5)
'''
np.random.rand(5, 5) генерирует массив размером 5x5,
где каждый элемент случайным образом выбирается из равномерного распределения от 0 до 1.
'''

'''
Преобразуем эту матрицу в объект DataFrame библиотеки pandas,
указав названия столбцов как 'Column_i', где i - индекс столбца:
'''
df = pd.DataFrame(data, columns=[f'Column_{i+1}' for i in range(data.shape[1])])
'''
pd.DataFrame(data, columns=[f'Column_{i+1}' for i in range(data.shape[1])]) создает DataFrame из
массива data, где каждый столбец будет иметь название вида 'Column_i', где i - порядковый номер столбца.
'''

'''
Сохраняем полученные данные в файл Excel. 
Создаем Excel-файл с рандомным именем random_data.xlsx и сохраняем в него данные DataFrame без индексов:
'''
excel_file = 'random_data.xlsx'
df.to_excel(excel_file, index=False)
'''
df.to_excel(excel_file, index=False) сохраняет DataFrame df в
файл Excel с именем random_data.xlsx, не включая индексы строк.
'''

'''
Считываем данные из Excel-файла обратно в объект DataFrame:
'''
df_from_excel = pd.read_excel(excel_file)
'''
pd.read_excel(excel_file) считывает данные из файла Excel random_data.xlsx и загружает их в DataFrame df_from_excel.
'''

'''
Преобразуем DataFrame обратно в массив NumPy:
'''
numpy_array = df_from_excel.to_numpy()
'''
df_from_excel.to_numpy() преобразует DataFrame df_from_excel в массив NumPy.
'''

'''
Выводим первые 5 строк полученного массива NumPy:
'''
print(numpy_array[:5])
'''
print(numpy_array[:5]) выводит первые 5 строк массива NumPy numpy_array.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Вариант №2.
'''
import numpy as np
import pandas as pd

# Снова создаем генерацию в которой снова будет рандомная матрица данных
data = np.random.rand(5, 5)

# Снова создаем DataFrame из матрицы данных как и в первом варианте
df = pd.DataFrame(data)

# Далее сохраняем данные в Excel-файл
excel_file = 'random_data.xlsx'  # Имя Excel-файла
df.to_excel(excel_file, index=False)  # Сохраняем данные без индексов

# Теперь по логике вещей должна происходить загрузка данных из Excel-файла в NumPy массив
loaded_data = pd.read_excel(excel_file, header=None).to_numpy()

# По итогу Мы видим вывод первых нескольких строк массива для проверки
print(loaded_data[:5])  # Печать первых 5 строк массива (можно изменить число строк)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Как можно заметить, этот код выполняет то же самое, что и предыдущий,
но использует немного другой синтаксис для генерации данных, создания DataFrame и сохранения в Excel-файл.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Импортируем библиотеки numpy и pandas:
'''
import numpy as np
import pandas as pd
'''
numpy используется для работы с многомерными массивами, а pandas - для работы с данными в виде таблиц (DataFrame).
'''

'''
Создаем случайную матрицу размером 5x5, 
заполненную случайными значениями от 0 до 1 с помощью функции numpy.random.rand():
'''
data = np.random.rand(5, 5)
'''
np.random.rand(5, 5) генерирует массив размером 5x5,
где каждый элемент случайным образом выбирается из равномерного распределения от 0 до 1.
'''

'''
Создаем DataFrame из этой матрицы данных:
'''
df = pd.DataFrame(data)
'''
pd.DataFrame(data) создает DataFrame df из массива data.
'''

'''
Сохраняем полученные данные в файл Excel. 
Создаем Excel-файл с именем random_data.xlsx и сохраняем в него данные DataFrame без индексов:
'''
excel_file = 'random_data.xlsx'
df.to_excel(excel_file, index=False)
'''
df.to_excel(excel_file, index=False) сохраняет DataFrame df в файл Excel с именем random_data.xlsx,
не включая индексы строк.
'''

'''
Загружаем данные из Excel-файла обратно в массив NumPy:
'''
loaded_data = pd.read_excel(excel_file, header=None).to_numpy()
'''
pd.read_excel(excel_file, header=None) считывает данные из файла Excel random_data.xlsx, а header=None указывает на то,
что в файле нет строки заголовка. Затем метод to_numpy() преобразует DataFrame в массив NumPy.
'''

'''
Выводим первые 5 строк полученного массива NumPy:
'''
print(loaded_data[:5])
'''
print(loaded_data[:5]) выводит первые 5 строк массива loaded_data.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Вариант №3.
'''
import numpy as np
import pandas as pd
import openpyxl

# Снова генерация рандомной матрицы данных
data = np.random.rand(50, 50)

# Создание DataFrame из матрицы данных
df = pd.DataFrame(data)

# Сохранение данных в Excel-файл
excel_file = 'random_data.xlsx'  # Имя Excel-файла
try:
    with pd.ExcelWriter(excel_file) as writer:
        df.to_excel(writer, index=False)  # Сохраняем данные без индексов
except Exception as e:
    print(f"Ошибка при сохранении в файл Excel: {e}")

# Чтение данных из Excel-файла с помощью openpyxl
try:
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    loaded_data = np.array([[cell.value for cell in row] for row in sheet.iter_rows()])
    print("Загруженные данные из файла Excel:")
    print(loaded_data)
except Exception as e:
    print(f"Ошибка при чтении данных из файла Excel: {e}")

# Вывод первых нескольких строк массива для проверки
if loaded_data.size > 0:
    print("Первые 25 строк загруженных данных:")
    print(loaded_data[:25])  # Печать первых 5 строк массива (можно изменить число строк)
else:
    print("Данные не загружены из файла Excel.")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Импорт библиотек numpy, pandas и openpyxl:
'''
import numpy as np
import pandas as pd
import openpyxl
'''
numpy используется для работы с многомерными массивами,
pandas - для работы с данными в виде таблиц (DataFrame), а openpyxl - для работы с файлами Excel.
'''

'''
Создание случайной матрицы данных размером 50x50 с помощью функции numpy.random.rand():
'''
data = np.random.rand(50, 50)
'''
np.random.rand(50, 50) генерирует массив размером 50x50,
где каждый элемент случайным образом выбирается из равномерного распределения от 0 до 1.
'''

'''
Создание DataFrame из этой матрицы данных:
'''
df = pd.DataFrame(data)
'''
pd.DataFrame(data) создает DataFrame df из массива data.
'''

'''
Сохранение полученных данных в файл Excel с помощью pd.ExcelWriter():
'''
excel_file = 'random_data.xlsx'
try:
    with pd.ExcelWriter(excel_file) as writer:
        df.to_excel(writer, index=False)
except Exception as e:
    print(f"Ошибка при сохранении в файл Excel: {e}")
'''
pd.ExcelWriter(excel_file) создает объект для записи в файл Excel.
df.to_excel(writer, index=False) сохраняет данные DataFrame df в файл Excel без индексов.
'''

'''
Чтение данных из созданного Excel-файла с помощью библиотеки openpyxl:
'''
try:
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    loaded_data = np.array([[cell.value for cell in row] for row in sheet.iter_rows()])
    print("Загруженные данные из файла Excel:")
    print(loaded_data)
except Exception as e:
    print(f"Ошибка при чтении данных из файла Excel: {e}")
'''
openpyxl.load_workbook(excel_file) загружает файл Excel random_data.xlsx.
sheet.iter_rows() перебирает строки в листе Excel.
loaded_data = np.array([[cell.value for cell in row] for row in sheet.iter_rows()]) создает
двумерный массив из данных в листе Excel.
'''

'''
Вывод первых нескольких строк загруженного массива данных:
'''
if loaded_data.size > 0:
    print("Первые 25 строк загруженных данных:")
    print(loaded_data[:25])  # Печать первых 25 строк массива (можно изменить число строк)
else:
    print("Данные не загружены из файла Excel.")
'''
Если загруженные данные не пусты, выводятся первые 25 строк массива loaded_data.
Если массив пустой, выводится сообщение о том, что данные не загружены.
'''
